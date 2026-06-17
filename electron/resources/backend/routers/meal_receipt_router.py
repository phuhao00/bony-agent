"""
公司定制 · 每日餐费票据
========================
- SQLite 持久化：每位员工每天一条记录（UNIQUE(employee_id, meal_date)）
- 图片识别：复用现有 LLM 视觉模型（core.llm_provider，可运行时切换 provider）
- 个人 / 全员统计 + openpyxl 导出 Excel

存储：
- DB:    storage/meal/meal_receipts.db
- 图片:  storage/uploads/meal/<employee_id>/<uuid>.<ext>
"""
from __future__ import annotations

import base64
import hashlib
import hmac
import io
import json
import os
import re
import sqlite3
import time
import uuid
from datetime import date, datetime
from pathlib import Path
from typing import Any, Optional

MAX_UPLOAD_IMAGES = 3

from utils.logger import setup_logger

logger = setup_logger("meal_receipt")

PROJECT_ROOT = Path(__file__).parent.parent.parent


def _meal_storage_dir() -> Path:
    from services.meal_feishu_config import get_storage_dir

    return get_storage_dir() / "meal"


def _uploads_meal_dir() -> Path:
    from services.meal_feishu_config import get_storage_dir

    return get_storage_dir() / "uploads" / "meal"


def meal_dir() -> Path:
    """运行时解析，兼容 Electron STORAGE_DIR（勿在模块加载时缓存路径）。"""
    return _meal_storage_dir()


def db_path() -> Path:
    return meal_dir() / "meal_receipts.db"


def img_dir() -> Path:
    return _uploads_meal_dir()


_initialized = False

_TOKEN_TTL = 2 * 3600


# ──────────────────────────────────────────────
# 飞书员工身份令牌（H5 上传 / 深链，与 feishu_bot 算法一致）
# ──────────────────────────────────────────────
def _upload_secret() -> bytes:
    key = (
        os.getenv("FEISHU_APP_SECRET", "")
        or os.getenv("MEAL_UPLOAD_SECRET", "")
        or "meal-receipt-fallback-secret"
    )
    return key.encode()


def _b64e(raw: bytes) -> str:
    return base64.urlsafe_b64encode(raw).decode().rstrip("=")


def _b64d(s: str) -> bytes:
    pad = "=" * (-len(s) % 4)
    return base64.urlsafe_b64decode(s + pad)


def make_upload_token(open_id: str, name: str = "") -> str:
    payload = {"oid": open_id, "name": name, "exp": int(time.time()) + _TOKEN_TTL}
    body = _b64e(json.dumps(payload, ensure_ascii=False).encode())
    sig = _b64e(hmac.new(_upload_secret(), body.encode(), hashlib.sha256).digest())
    return f"{body}.{sig}"


def employee_id_from_name(name: str) -> str:
    """表单上传：用姓名生成稳定 employee_id（每人每天一条仍按此 id）。"""
    n = re.sub(r"\s+", "", (name or "").strip())
    if not n:
        return ""
    digest = hashlib.sha256(n.encode("utf-8")).hexdigest()[:16]
    return f"name:{digest}"


def employee_ids_for_lookup(employee_id: str = "", employee_name: str = "") -> list[str]:
    """合并同一人的多种 employee_id（ou_ / name: / 历史库内 id）。"""
    ids: set[str] = set()
    eid = (employee_id or "").strip()
    if eid:
        ids.add(eid)
    name = (employee_name or "").strip()
    if name:
        nid = employee_id_from_name(name)
        if nid:
            ids.add(nid)
        init_db()
        c = _conn()
        try:
            rows = c.execute(
                "SELECT DISTINCT employee_id FROM meal_receipts "
                "WHERE TRIM(employee_name)=?",
                (name,),
            ).fetchall()
            for row in rows:
                if row[0]:
                    ids.add(str(row[0]))
        finally:
            c.close()
    return sorted(ids)


def _record_rank(rec: dict) -> tuple[int, str]:
    eid = str(rec.get("employee_id") or "")
    pref = 2 if eid.startswith("ou_") else 1 if eid.startswith("name:") else 0
    ts = str(rec.get("updated_at") or rec.get("created_at") or "")
    return pref, ts


def _dedupe_by_meal_date(records: list[dict]) -> list[dict]:
    """同人同日可能因 id 分裂存了两条，保留更可信的一条（优先 ou_、较新 updated_at）。"""
    best: dict[str, dict] = {}
    for rec in records:
        d = rec.get("meal_date") or ""
        if not d:
            continue
        cur = best.get(d)
        if not cur or _record_rank(rec) > _record_rank(cur):
            best[d] = rec
    return sorted(best.values(), key=lambda x: x.get("meal_date") or "", reverse=True)


def find_existing_receipt(
    employee_id: str,
    employee_name: str,
    meal_date: str,
) -> Optional[dict]:
    for alt_id in employee_ids_for_lookup(employee_id, employee_name):
        rec = get_by_emp_date(alt_id, meal_date)
        if rec:
            return rec
    return None


def resolve_upload_identity(
    token: str = "",
    employee_name: str = "",
) -> tuple[str, str, str]:
    """根据上传 token 或姓名解析 employee_id。返回 (employee_id, display_name, error)。"""
    t = (token or "").strip()
    if t:
        info = verify_upload_token(t)
        if not info:
            return "", "", "链接无效或已过期，请在飞书重新发送「餐费」"
        oid = (info.get("oid") or "").strip()
        if not oid:
            return "", "", "令牌无效"
        name = (info.get("name") or info.get("nickname") or "").strip()
        return oid, name, ""
    name = (employee_name or "").strip()
    if name:
        eid = employee_id_from_name(name)
        if not eid:
            return "", "", "姓名无效"
        return eid, name, ""
    return "", "", "需要 token 或姓名"


def verify_upload_token(token: str) -> Optional[dict]:
    try:
        body, sig = token.split(".", 1)
        expected = _b64e(hmac.new(_upload_secret(), body.encode(), hashlib.sha256).digest())
        if not hmac.compare_digest(sig, expected):
            return None
        payload = json.loads(_b64d(body))
        if int(payload.get("exp", 0)) < int(time.time()):
            return None
        return {"oid": payload.get("oid", ""), "name": payload.get("name", "")}
    except Exception:
        return None


def feishu_integration_status() -> dict:
    """供 Web 展示飞书接入状态（lark-cli 或 lark-oapi 长连接）。"""
    from services.meal_feishu_config import load_config, is_configured, uses_lark_cli
    from services import meal_feishu_ws as fws

    cfg = load_config()
    app_id = cfg.get("app_id", "")
    from utils.meal_public_url import meal_public_web_base, meal_upload_page_url

    agent_web = meal_public_web_base()
    public = agent_web if agent_web != "http://localhost:3000" else ""
    out: dict = {
        "feishu_app_configured": is_configured(),
        "use_lark_cli": uses_lark_cli(),
        "connection_mode": fws.connection_mode(),
        "feishu_app_id_prefix": (app_id[:8] + "…") if len(app_id) > 8 else "",
        "ws_connected": fws.is_connected(),
        "ws_error": fws.last_error(),
        "public_base_url": public,
        "meal_web_base_url": agent_web,
        "upload_path": meal_upload_page_url(),
        "public_upload_url": meal_upload_page_url(),
        "webhook_path": "/meal/feishu/webhook",
        "lark_cli_page": f"{agent_web}/lark-cli",
        "commands": [
            "餐费", "餐费记录", "餐费统计", "餐费补录", "餐费删除",
            "私聊/群@后发餐费截图",
        ],
    }
    try:
        from services import meal_feishu_lark_cli as lc

        out.update(lc.integration_probe())
    except Exception as e:
        out["lark_cli_error"] = str(e)[:200]
    return out


def _norm_record(rec: Optional[dict]) -> Optional[dict]:
    """兼容飞书侧 user_id 字段名，并附加日封顶后报销金额。"""
    if not rec:
        return rec
    out = enrich_record(dict(rec))
    if out.get("employee_id") and not out.get("user_id"):
        out["user_id"] = out["employee_id"]
    out["image_urls"] = image_urls_for_record(out)
    out["pending_review"] = bool(out.get("pending_review"))
    return out


def image_urls_for_record(rec: dict) -> list[str]:
    """解析记录中的全部票据图片 URL（最多 3 张）。"""
    urls: list[str] = []
    raw = rec.get("image_urls")
    if raw:
        try:
            parsed = json.loads(raw) if isinstance(raw, str) else raw
            if isinstance(parsed, list):
                urls.extend(str(u).strip() for u in parsed if u)
        except Exception:
            pass
    primary = (rec.get("image_url") or "").strip()
    if primary and primary not in urls:
        urls.insert(0, primary)
    deduped: list[str] = []
    for u in urls:
        if u and u not in deduped:
            deduped.append(u)
    return deduped[:MAX_UPLOAD_IMAGES]


def _migrate_db_columns(c: sqlite3.Connection) -> None:
    cols = {row[1] for row in c.execute("PRAGMA table_info(meal_receipts)").fetchall()}
    if "pending_review" not in cols:
        c.execute("ALTER TABLE meal_receipts ADD COLUMN pending_review INTEGER DEFAULT 0")
    if "review_note" not in cols:
        c.execute("ALTER TABLE meal_receipts ADD COLUMN review_note TEXT DEFAULT ''")
    if "image_urls" not in cols:
        c.execute("ALTER TABLE meal_receipts ADD COLUMN image_urls TEXT DEFAULT ''")


# ──────────────────────────────────────────────
# DB 层
# ──────────────────────────────────────────────
def _now() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def _conn() -> sqlite3.Connection:
    meal_dir().mkdir(parents=True, exist_ok=True)
    c = sqlite3.connect(str(db_path()))
    c.row_factory = sqlite3.Row
    c.execute("PRAGMA journal_mode=WAL")
    return c


def init_db() -> None:
    global _initialized
    if _initialized:
        return
    c = _conn()
    try:
        c.execute(
            """
            CREATE TABLE IF NOT EXISTS meal_receipts (
                id            TEXT PRIMARY KEY,
                employee_id   TEXT NOT NULL,
                employee_name TEXT,
                meal_date     TEXT NOT NULL,
                amount        REAL NOT NULL,
                currency      TEXT DEFAULT 'CNY',
                merchant      TEXT,
                image_url     TEXT,
                raw_ocr       TEXT,
                source        TEXT,
                created_at    TEXT,
                updated_at    TEXT,
                UNIQUE(employee_id, meal_date)
            )
            """
        )
        c.execute("CREATE INDEX IF NOT EXISTS idx_meal_emp ON meal_receipts(employee_id)")
        c.execute("CREATE INDEX IF NOT EXISTS idx_meal_date ON meal_receipts(meal_date)")
        _migrate_db_columns(c)
        c.commit()
        _initialized = True
        logger.info(f"[meal] DB 初始化完成 | {db_path()}")
    finally:
        c.close()


def get_by_emp_date(employee_id: str, meal_date: str) -> Optional[dict]:
    init_db()
    c = _conn()
    try:
        row = c.execute(
            "SELECT * FROM meal_receipts WHERE employee_id=? AND meal_date=?",
            (employee_id, meal_date),
        ).fetchone()
        return dict(row) if row else None
    finally:
        c.close()


def upsert_receipt(
    *,
    employee_id: str,
    employee_name: str,
    meal_date: str,
    amount: float,
    currency: str = "CNY",
    merchant: str = "",
    image_url: str = "",
    image_urls: str = "",
    raw_ocr: str = "",
    source: str = "web_upload",
    overwrite: bool = False,
    pending_review: bool = False,
    review_note: str = "",
) -> tuple[bool, str, Optional[dict]]:
    """返回 (ok, status, record)；status: created|updated|exists。"""
    init_db()
    existing = find_existing_receipt(employee_id, employee_name, meal_date)
    if existing and not overwrite:
        return False, "exists", existing

    row_eid = employee_id
    if existing:
        row_eid = str(existing.get("employee_id") or employee_id)
        if employee_id.startswith("ou_") and not row_eid.startswith("ou_"):
            row_eid = employee_id

    c = _conn()
    try:
        now = _now()
        if existing:
            c.execute(
                """
                UPDATE meal_receipts
                SET employee_id=?, employee_name=?, amount=?, currency=?, merchant=?,
                    image_url=?, image_urls=?, raw_ocr=?, source=?, updated_at=?,
                    pending_review=?, review_note=?
                WHERE employee_id=? AND meal_date=?
                """,
                (
                    row_eid, employee_name, amount, currency, merchant,
                    image_url or existing.get("image_url", ""),
                    image_urls or existing.get("image_urls", ""),
                    raw_ocr, source, now,
                    1 if pending_review else 0, (review_note or "")[:500],
                    existing["employee_id"], meal_date,
                ),
            )
            c.commit()
            return True, "updated", get_by_emp_date(row_eid, meal_date)

        c.execute(
            """
            INSERT INTO meal_receipts
                (id, employee_id, employee_name, meal_date, amount, currency,
                 merchant, image_url, image_urls, raw_ocr, source, created_at, updated_at,
                 pending_review, review_note)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """,
            (
                str(uuid.uuid4()), employee_id, employee_name, meal_date, amount,
                currency, merchant, image_url, image_urls, raw_ocr, source, now, now,
                1 if pending_review else 0, (review_note or "")[:500],
            ),
        )
        c.commit()
        return True, "created", get_by_emp_date(employee_id, meal_date)
    finally:
        c.close()


def _month_clause(month: str) -> tuple[str, tuple]:
    if month:
        return " AND meal_date LIKE ?", (f"{month}-%",)
    return "", tuple()


def list_by_emp(
    employee_id: str,
    month: str = "",
    *,
    employee_name: str = "",
) -> list[dict]:
    ids = employee_ids_for_lookup(employee_id, employee_name)
    if not ids:
        return []
    init_db()
    clause, params = _month_clause(month)
    placeholders = ",".join("?" * len(ids))
    c = _conn()
    try:
        rows = c.execute(
            f"SELECT * FROM meal_receipts WHERE employee_id IN ({placeholders})"
            f"{clause} ORDER BY meal_date DESC, updated_at DESC",
            (*ids, *params),
        ).fetchall()
        return _dedupe_by_meal_date([dict(r) for r in rows])
    finally:
        c.close()


def list_all(month: str = "") -> list[dict]:
    init_db()
    clause, params = _month_clause(month)
    c = _conn()
    try:
        rows = c.execute(
            f"SELECT * FROM meal_receipts WHERE 1=1{clause} "
            f"ORDER BY meal_date DESC, employee_name ASC",
            tuple(params),
        ).fetchall()
        return [dict(r) for r in rows]
    finally:
        c.close()


def delete_receipt(employee_id: str, meal_date: str) -> bool:
    init_db()
    c = _conn()
    try:
        cur = c.execute(
            "DELETE FROM meal_receipts WHERE employee_id=? AND meal_date=?",
            (employee_id, meal_date),
        )
        c.commit()
        return cur.rowcount > 0
    finally:
        c.close()


# ──────────────────────────────────────────────
# 视觉识别（复用 core.llm_provider，可运行时切换 provider）
# ──────────────────────────────────────────────
_VISION_PROMPT = (
    "你是餐费票据/小票识别助手。查看图片并提取就餐消费信息。"
    "只输出一个 JSON 对象，不要解释、不要 markdown 代码块。字段：\n"
    '- date: 消费日期 YYYY-MM-DD；只有月日则补当前年份；无法判断填 null\n'
    "- amount: 实付/合计金额数字（人民币元，两位小数）；识别不到填 null\n"
    '- currency: 币种，默认 "CNY"\n'
    '- merchant: 商家名称，识别不到填 ""\n'
    "- confidence: 0~1 置信度\n"
    "若不是餐费票据，amount 返回 null。"
)

_MIME = {
    ".jpg": "image/jpeg", ".jpeg": "image/jpeg", ".png": "image/png",
    ".webp": "image/webp", ".gif": "image/gif", ".bmp": "image/bmp",
}


def get_vision_info() -> dict:
    try:
        from core.llm_provider import (
            get_provider_id, get_api_key, get_vision_model,
            is_vision_capable_model, default_vision_model_for_provider,
        )
        provider_id = get_provider_id()
        model = get_vision_model(provider_id)
        if not is_vision_capable_model(model):
            model = default_vision_model_for_provider(provider_id)
        return {
            "provider": provider_id,
            "model": model,
            "configured": bool(get_api_key(provider_id)),
        }
    except Exception as e:
        return {"provider": "", "model": "", "configured": False, "error": str(e)}


def _parse_json(text: str) -> dict:
    if not text:
        return {}
    s = text.strip()
    s = re.sub(r"^```(?:json)?\s*", "", s)
    s = re.sub(r"\s*```$", "", s)
    try:
        return json.loads(s)
    except Exception:
        m = re.search(r"\{.*\}", s, re.DOTALL)
        if m:
            try:
                return json.loads(m.group(0))
            except Exception:
                return {}
    return {}


def _norm_date(raw) -> Optional[str]:
    if not raw or not isinstance(raw, str):
        return None
    s = raw.strip().replace("/", "-").replace(".", "-")
    s = s.replace("年", "-").replace("月", "-").replace("日", "").strip("-")
    m = re.match(r"^(\d{4})-(\d{1,2})-(\d{1,2})$", s)
    if m:
        y, mo, d = m.groups()
        return f"{int(y):04d}-{int(mo):02d}-{int(d):02d}"
    m2 = re.match(r"^(\d{1,2})-(\d{1,2})$", s)
    if m2:
        mo, d = m2.groups()
        return f"{date.today().year:04d}-{int(mo):02d}-{int(d):02d}"
    return None


def _norm_amount(raw) -> Optional[float]:
    if raw is None:
        return None
    try:
        if isinstance(raw, str):
            raw = re.sub(r"[^\d.]", "", raw)
            if not raw:
                return None
        v = round(float(raw), 2)
        return v if v > 0 else None
    except Exception:
        return None


def recognize_receipt(image_path: str) -> dict:
    """识别餐费票据，返回 {ok, date, amount, currency, merchant, confidence, model, raw, error}"""
    from openai import OpenAI
    from core.llm_provider import (
        get_provider_id, get_api_key, get_provider_config,
        get_vision_model, is_vision_capable_model, default_vision_model_for_provider,
    )

    provider_id = get_provider_id()
    key = get_api_key(provider_id)
    if not key:
        return {"ok": False, "error": "视觉模型 API Key 未配置（请在「设置·模型」中配置当前 provider 的 Key）"}

    p = Path(image_path)
    if not p.exists():
        return {"ok": False, "error": f"图片不存在：{image_path}"}

    try:
        cfg = get_provider_config()
        model = get_vision_model(provider_id)
        if not is_vision_capable_model(model):
            model = default_vision_model_for_provider(provider_id)

        mime = _MIME.get(p.suffix.lower(), "image/jpeg")
        b64 = base64.b64encode(p.read_bytes()).decode()
        data_uri = f"data:{mime};base64,{b64}"

        client = OpenAI(api_key=key, base_url=cfg.base_url)
        resp = client.chat.completions.create(
            model=model,
            messages=[{
                "role": "user",
                "content": [
                    {"type": "image_url", "image_url": {"url": data_uri}},
                    {"type": "text", "text": _VISION_PROMPT},
                ],
            }],
            max_tokens=500,
            temperature=0,
        )
        raw = resp.choices[0].message.content or ""
        logger.info(f"[meal] vision provider={provider_id} model={model} raw={raw[:120]!r}")

        parsed = _parse_json(raw)
        amount = _norm_amount(parsed.get("amount"))
        if amount is None:
            return {
                "ok": False,
                "error": "未识别到有效金额，请确认是餐费票据，或手动填写金额",
                "raw": raw, "model": model,
            }
        return {
            "ok": True,
            "date": _norm_date(parsed.get("date")),
            "amount": amount,
            "currency": (parsed.get("currency") or "CNY").upper()[:8],
            "merchant": (parsed.get("merchant") or "").strip()[:100],
            "confidence": parsed.get("confidence"),
            "model": model,
            "raw": raw,
        }
    except Exception as e:
        logger.error(f"[meal] 识别失败: {e}")
        return {"ok": False, "error": f"识别服务调用失败：{str(e)[:120]}"}


# ──────────────────────────────────────────────
# 图片保存
# ──────────────────────────────────────────────
def _per_image_rows(recognitions: list[dict]) -> list[dict]:
    rows: list[dict] = []
    for i, rec in enumerate(recognitions):
        rows.append(
            {
                "index": i + 1,
                "ok": bool(rec.get("ok")),
                "date": rec.get("date"),
                "amount": rec.get("amount"),
                "currency": rec.get("currency"),
                "merchant": rec.get("merchant"),
                "error": rec.get("error"),
                "model": rec.get("model"),
            }
        )
    return rows


def aggregate_upload_recognitions(
    recognitions: list[dict],
    *,
    meal_date_override: str = "",
) -> tuple[float, str, str, str, bool, str, dict[str, Any]]:
    """
    合并多张票据识别结果。
    返回: amount, currency, meal_date, merchant, pending_review, review_note, recognized_payload
    """
    today = date.today().strftime("%Y-%m-%d")
    explicit = bool((meal_date_override or "").strip())
    md = (meal_date_override or "").strip() or today

    amounts: list[float] = []
    merchants: list[str] = []
    ocr_dates: list[str] = []
    per_image = _per_image_rows(recognitions)
    currency = "CNY"

    for rec in recognitions:
        if not rec.get("ok"):
            continue
        amt = _norm_amount(rec.get("amount"))
        if amt is not None:
            amounts.append(amt)
        m = (rec.get("merchant") or "").strip()
        if m and m not in merchants:
            merchants.append(m)
        d = _norm_date(rec.get("date")) if rec.get("date") else None
        if d:
            ocr_dates.append(d)
        currency = (rec.get("currency") or currency or "CNY").upper()[:8]

    if not amounts:
        return (
            0.0,
            currency,
            md,
            "",
            True,
            "未识别到有效金额",
            {"images": per_image, "image_count": len(recognitions)},
        )

    if not explicit:
        if ocr_dates:
            md = ocr_dates[0]
        else:
            md = today

    notes: list[str] = []
    pending = False
    if not ocr_dates:
        pending = True
        notes.append("未识别到日期")
    else:
        unique_dates = sorted(set(ocr_dates))
        if len(unique_dates) > 1:
            pending = True
            notes.append("多张票据日期不一致")
        for d in ocr_dates:
            if d != md:
                pending = True
                notes.append(f"票据日期{d}与登记日{md}不符")
                break

    total = round(sum(amounts), 2)
    merchant = "，".join(merchants)
    payload = {
        "images": per_image,
        "image_count": len(recognitions),
        "amount_total": total,
        "amounts": amounts,
        "meal_date": md,
        "merchant": merchant,
        "pending_review": pending,
        "review_note": "；".join(notes),
    }
    return total, currency, md, merchant, pending, "；".join(notes), payload


def save_receipt_image(employee_id: str, content: bytes, filename: str) -> tuple[str, str]:
    """保存票据图片，返回 (本地绝对路径, 可访问 url /uploads/meal/...)。"""
    ext = ("." + filename.rsplit(".", 1)[-1].lower()) if "." in filename else ".jpg"
    if ext not in _MIME:
        ext = ".jpg"
    safe_emp = re.sub(r"[^\w\-]", "_", employee_id)[:64] or "anon"
    d = img_dir() / safe_emp
    d.mkdir(parents=True, exist_ok=True)
    name = f"{uuid.uuid4().hex}{ext}"
    fpath = d / name
    fpath.write_bytes(content)
    return str(fpath), f"/uploads/meal/{safe_emp}/{name}"


def file_content_sha256(content: bytes) -> str:
    return hashlib.sha256(content).hexdigest()


def dedupe_upload_files(
    files: list[tuple[bytes, str]],
) -> tuple[list[tuple[bytes, str]], int]:
    """单次上传内按文件内容去重。"""
    seen: set[str] = set()
    out: list[tuple[bytes, str]] = []
    skipped = 0
    for content, filename in files:
        digest = file_content_sha256(content)
        if digest in seen:
            skipped += 1
            continue
        seen.add(digest)
        out.append((content, filename))
    return out, skipped


def image_hashes_from_record(record: Optional[dict]) -> set[str]:
    """当日已有凭证图片的内容哈希，用于跨次上传去重。"""
    if not record:
        return set()
    hashes: set[str] = set()
    for url in image_urls_for_record(record):
        path = resolve_receipt_image_path(url)
        if not path:
            continue
        try:
            hashes.add(file_content_sha256(path.read_bytes()))
        except Exception:
            pass
    return hashes


def filter_files_not_in_hashes(
    files: list[tuple[bytes, str]],
    existing_hashes: set[str],
) -> tuple[list[tuple[bytes, str]], int]:
    """去掉与已有凭证内容相同的上传文件。"""
    if not existing_hashes:
        return files, 0
    out: list[tuple[bytes, str]] = []
    skipped = 0
    for content, filename in files:
        if file_content_sha256(content) in existing_hashes:
            skipped += 1
            continue
        out.append((content, filename))
    return out, skipped


def recognitions_from_raw_ocr(raw_ocr: str) -> list[dict]:
    """从已入库 raw_ocr 恢复各张图的识别结果，供合并重算金额。"""
    if not raw_ocr:
        return []
    try:
        payload = json.loads(raw_ocr) if isinstance(raw_ocr, str) else raw_ocr
    except Exception:
        return []
    if not isinstance(payload, dict):
        return []
    images = payload.get("images")
    if not isinstance(images, list):
        return []
    recs: list[dict] = []
    for row in images:
        if not isinstance(row, dict) or not row.get("ok"):
            continue
        recs.append({
            "ok": True,
            "date": row.get("date"),
            "amount": row.get("amount"),
            "currency": row.get("currency") or "CNY",
            "merchant": row.get("merchant") or "",
        })
    return recs


def merge_image_url_lists(
    existing_record: Optional[dict],
    new_urls: list[str],
) -> list[str]:
    """合并当日图片 URL（URL 去重、最多 MAX_UPLOAD_IMAGES 张）。"""
    combined: list[str] = []
    if existing_record:
        combined.extend(image_urls_for_record(existing_record))
    for url in new_urls:
        u = (url or "").strip()
        if u and u not in combined:
            combined.append(u)
    return combined[:MAX_UPLOAD_IMAGES]


def merge_saved_images_into_record(
    existing_record: Optional[dict],
    saved: list[tuple[str, str]],
) -> list[str]:
    """按文件内容合并新保存的图片，跳过与已有凭证相同的文件。"""
    if not saved:
        return image_urls_for_record(existing_record) if existing_record else []
    seen = image_hashes_from_record(existing_record)
    new_urls: list[str] = []
    for img_path, img_url in saved:
        try:
            digest = file_content_sha256(Path(img_path).read_bytes())
        except Exception:
            digest = ""
        if digest and digest in seen:
            continue
        if digest:
            seen.add(digest)
        new_urls.append(img_url)
    return merge_image_url_lists(existing_record, new_urls)


# ──────────────────────────────────────────────
# 统计（每日报销封顶）
# ──────────────────────────────────────────────
DAILY_REIMBURSEMENT_CAP = float(os.getenv("MEAL_DAILY_CAP", "30"))


def reimbursement_amount(amount: float | int | str) -> float:
    """单日可报销金额：不超过日封顶，低于封顶按实报。"""
    try:
        raw = float(amount)
    except (TypeError, ValueError):
        raw = 0.0
    return round(min(max(raw, 0.0), DAILY_REIMBURSEMENT_CAP), 2)


def enrich_record(record: dict) -> dict:
    """为单条记录附加报销金额（日封顶后）。"""
    out = dict(record)
    bill = float(out.get("amount") or 0)
    reimb = reimbursement_amount(bill)
    out["reimbursement_amount"] = reimb
    out["daily_cap"] = DAILY_REIMBURSEMENT_CAP
    out["capped"] = bill > DAILY_REIMBURSEMENT_CAP
    return out


def enrich_records(
    records: list[dict],
    *,
    with_attendance: bool = True,
    month: str = "",
) -> list[dict]:
    base = [enrich_record(r) for r in records]
    if not with_attendance:
        return base
    try:
        from services.meal_feishu_attendance import attach_attendance

        return attach_attendance(base, month=month)
    except Exception as e:
        logger.warning("[meal] 附加考勤失败: %s", e)
        return base


def summarize(records: list[dict]) -> dict:
    """汇总：total 为可报销合计，total_bill 为票据原始合计。"""
    if not records:
        return {
            "total": 0.0,
            "total_bill": 0.0,
            "days": 0,
            "avg": 0.0,
            "count": 0,
            "daily_cap": DAILY_REIMBURSEMENT_CAP,
            "capped_days": 0,
        }
    bill = sum(float(r["amount"]) for r in records)
    reimb = sum(reimbursement_amount(r["amount"]) for r in records)
    days = len({r["meal_date"] for r in records})
    capped_days = sum(1 for r in records if float(r["amount"]) > DAILY_REIMBURSEMENT_CAP)
    return {
        "total": round(reimb, 2),
        "total_bill": round(bill, 2),
        "days": days,
        "avg": round(reimb / days, 2) if days else 0.0,
        "count": len(records),
        "daily_cap": DAILY_REIMBURSEMENT_CAP,
        "capped_days": capped_days,
    }


def summarize_by_user(records: list[dict]) -> list[dict]:
    g: dict[str, dict] = {}
    for r in records:
        e = g.setdefault(r["employee_id"], {
            "employee_id": r["employee_id"],
            "employee_name": r.get("employee_name") or r["employee_id"],
            "total": 0.0,
            "total_bill": 0.0,
            "dates": set(),
            "capped_days": 0,
        })
        bill = float(r["amount"])
        e["total_bill"] += bill
        e["total"] += reimbursement_amount(bill)
        e["dates"].add(r["meal_date"])
        if bill > DAILY_REIMBURSEMENT_CAP:
            e["capped_days"] += 1
        e["employee_name"] = r.get("employee_name") or e["employee_name"]
    out = []
    for e in g.values():
        days = len(e["dates"])
        out.append({
            "employee_id": e["employee_id"],
            "employee_name": e["employee_name"],
            "total": round(e["total"], 2),
            "total_bill": round(e["total_bill"], 2),
            "days": days,
            "avg": round(e["total"] / days, 2) if days else 0.0,
            "daily_cap": DAILY_REIMBURSEMENT_CAP,
            "capped_days": e["capped_days"],
        })
    out.sort(key=lambda x: -x["total"])
    return out


def summarize_by_date(records: list[dict]) -> list[dict]:
    g: dict[str, dict] = {}
    for r in records:
        e = g.setdefault(r["meal_date"], {"total": 0.0, "total_bill": 0.0, "people": set()})
        bill = float(r["amount"])
        e["total_bill"] += bill
        e["total"] += reimbursement_amount(bill)
        e["people"].add(r["employee_id"])
    return [
        {
            "date": d,
            "total": round(v["total"], 2),
            "total_bill": round(v["total_bill"], 2),
            "people": len(v["people"]),
            "daily_cap": DAILY_REIMBURSEMENT_CAP,
        }
        for d, v in sorted(g.items(), reverse=True)
    ]


def current_month() -> str:
    return date.today().strftime("%Y-%m")


# ──────────────────────────────────────────────
# Excel 导出
# ──────────────────────────────────────────────
_EXCEL_IMG_MAX_W = 160
_EXCEL_IMG_MAX_H = 120
_EXCEL_IMG_ROW_HEIGHT = 95


def resolve_receipt_image_path(image_url: str) -> Optional[Path]:
    """将库内 image_url（/uploads/meal/...）解析为 storage 下的本地文件。"""
    u = (image_url or "").strip().replace("\\", "/")
    if not u:
        return None
    rel: Optional[str] = None
    if u.startswith("/uploads/meal/"):
        rel = u[len("/uploads/meal/") :].lstrip("/")
    elif "uploads/meal/" in u:
        rel = u.split("uploads/meal/", 1)[-1].lstrip("/")
    elif u.startswith(str(img_dir())):
        candidate = Path(u)
        if candidate.is_file():
            try:
                candidate.resolve().relative_to(img_dir().resolve())
                return candidate
            except ValueError:
                return None
    if not rel:
        return None
    try:
        candidate = (img_dir() / rel).resolve()
        candidate.relative_to(img_dir().resolve())
    except ValueError:
        return None
    return candidate if candidate.is_file() else None


def _embed_receipt_images(ws, records: list[dict], *, image_col: int, start_row: int = 2) -> None:
    """在明细表指定列嵌入员工上传的票据缩略图（最多 3 张横向排列）。"""
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.utils import get_column_letter

    thumb_w = max(48, _EXCEL_IMG_MAX_W // 3)
    for i, rec in enumerate(records):
        row = start_row + i
        urls = image_urls_for_record(rec)
        paths = [resolve_receipt_image_path(u) for u in urls]
        paths = [p for p in paths if p]
        if not paths:
            ws.cell(row=row, column=image_col, value="无图")
            ws.row_dimensions[row].height = 18
            continue
        placed = 0
        for j, path in enumerate(paths[:MAX_UPLOAD_IMAGES]):
            col = image_col + j
            anchor = f"{get_column_letter(col)}{row}"
            try:
                img = XLImage(str(path))
                w, h = float(img.width or thumb_w), float(img.height or _EXCEL_IMG_MAX_H)
                if w > 0 and h > 0:
                    scale = min(thumb_w / w, _EXCEL_IMG_MAX_H / h, 1.0)
                    img.width = int(w * scale)
                    img.height = int(h * scale)
                else:
                    img.width, img.height = thumb_w, _EXCEL_IMG_MAX_H
                ws.add_image(img, anchor)
                placed += 1
            except Exception as e:
                logger.warning(f"[meal] Excel 嵌入图片失败 {path}: {e}")
        if placed:
            ws.row_dimensions[row].height = _EXCEL_IMG_ROW_HEIGHT
        else:
            ws.cell(row=row, column=image_col, value="图片不可用")
            ws.row_dimensions[row].height = 18


def _styles():
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    return {
        "hf": Font(bold=True, color="FFFFFF"),
        "hfill": PatternFill("solid", fgColor="4472C4"),
        "c": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "b": Border(*[Side(style="thin", color="BFBFBF")] * 4),
    }


def _sheet(ws, headers, widths, rows):
    from openpyxl.utils import get_column_letter
    st = _styles()
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = st["hf"]; cell.fill = st["hfill"]; cell.alignment = st["c"]; cell.border = st["b"]
        ws.column_dimensions[get_column_letter(col)].width = w
    for i, row in enumerate(rows, 2):
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.alignment = st["c"]; cell.border = st["b"]


def build_personal_excel(records: list[dict], name: str, month: str) -> bytes:
    from openpyxl import Workbook

    enriched = enrich_records(records, month=month)
    wb = Workbook()
    ws = wb.active
    ws.title = "餐费明细"
    cap = DAILY_REIMBURSEMENT_CAP
    detail_rows = sorted(enriched, key=lambda x: x["meal_date"])
    rows = [
        [
            r["meal_date"],
            r["amount"],
            r["reimbursement_amount"],
            "是" if r.get("capped") else "否",
            r.get("currency", "CNY"),
            r.get("merchant", ""),
            "待处理" if r.get("pending_review") else "正常",
            r.get("review_note", "") if r.get("pending_review") else "",
            r.get("source", ""),
            r.get("clock_in") or "—",
            r.get("clock_out") or "—",
            r.get("updated_at", ""),
        ]
        for r in detail_rows
    ]
    img_col = 13
    _sheet(
        ws,
        [
            "日期",
            "票据金额",
            f"报销金额(日封顶{cap:g})",
            "是否封顶",
            "币种",
            "商家",
            "状态",
            "待处理说明",
            "来源",
            "上班打卡",
            "下班打卡",
            "更新时间",
            "票据截图",
        ],
        [14, 12, 16, 10, 8, 28, 10, 22, 14, 10, 10, 20, 22],
        rows,
    )
    _embed_receipt_images(ws, detail_rows, image_col=img_col)
    s = summarize(records)
    ws2 = wb.create_sheet("汇总")
    _sheet(
        ws2,
        ["员工", "月份", "日封顶", "可报销合计", "票据合计", "报销天数", "超标天数", "日均报销"],
        [16, 12, 10, 14, 14, 12, 12, 12],
        [
            [
                name,
                month or "全部",
                cap,
                s["total"],
                s["total_bill"],
                s["days"],
                s["capped_days"],
                s["avg"],
            ]
        ],
    )
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_company_excel(records: list[dict], month: str) -> bytes:
    from openpyxl import Workbook

    enriched = enrich_records(records, month=month)
    cap = DAILY_REIMBURSEMENT_CAP
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "全员明细"
    company_detail = sorted(
        enriched,
        key=lambda x: (x["meal_date"], x.get("employee_name") or ""),
    )
    img_col = 13
    _sheet(
        ws1,
        [
            "员工",
            "日期",
            "票据金额",
            f"报销金额(日封顶{cap:g})",
            "是否封顶",
            "币种",
            "商家",
            "状态",
            "待处理说明",
            "来源",
            "上班打卡",
            "下班打卡",
            "票据截图",
        ],
        [16, 14, 12, 16, 10, 8, 28, 10, 22, 14, 10, 10, 22],
        [
            [
                r.get("employee_name") or r["employee_id"],
                r["meal_date"],
                r["amount"],
                r["reimbursement_amount"],
                "是" if r.get("capped") else "否",
                r.get("currency", "CNY"),
                r.get("merchant", ""),
                "待处理" if r.get("pending_review") else "正常",
                r.get("review_note", "") if r.get("pending_review") else "",
                r.get("source", ""),
                r.get("clock_in") or "—",
                r.get("clock_out") or "—",
            ]
            for r in company_detail
        ],
    )
    _embed_receipt_images(ws1, company_detail, image_col=img_col)
    ws2 = wb.create_sheet("按员工汇总")
    _sheet(
        ws2,
        ["员工", "可报销合计", "票据合计", "报销天数", "超标天数", "日均报销"],
        [18, 14, 14, 12, 12, 12],
        [
            [
                u["employee_name"],
                u["total"],
                u["total_bill"],
                u["days"],
                u["capped_days"],
                u["avg"],
            ]
            for u in summarize_by_user(records)
        ],
    )
    ws3 = wb.create_sheet("按日汇总")
    _sheet(
        ws3,
        ["日期", "可报销合计", "票据合计", "报销人数"],
        [14, 14, 14, 12],
        [
            [d["date"], d["total"], d["total_bill"], d["people"]]
            for d in sorted(summarize_by_date(records), key=lambda x: x["date"])
        ],
    )
    s = summarize(records)
    ws4 = wb.create_sheet("全员汇总")
    _sheet(
        ws4,
        ["月份", "日封顶", "可报销合计", "票据合计", "记录条数", "超标天数"],
        [12, 10, 14, 14, 12, 12],
        [[month or "全部", cap, s["total"], s["total_bill"], s["count"], s["capped_days"]]],
    )
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ──────────────────────────────────────────────
# 高层管道（供 API 调用）
# ──────────────────────────────────────────────
def process_upload_feishu(
    *, token: str, content: bytes, filename: str, overwrite: bool = False,
    files: Optional[list[tuple[bytes, str]]] = None,
    manual_amount: Optional[float] = None,
) -> dict:
    """飞书 H5：用 token 识别员工后上传。"""
    info = verify_upload_token(token)
    if not info or not info.get("oid"):
        return {"ok": False, "status": "error", "error": "链接无效或已过期，请在飞书重新发送「餐费」获取上传入口"}
    upload_files = files if files else [(content, filename)]
    result = process_upload(
        employee_id=info["oid"],
        employee_name=info.get("name") or "",
        files=upload_files,
        overwrite=overwrite,
        source="feishu_web",
        manual_amount=manual_amount,
    )
    if result.get("record"):
        result["record"] = _norm_record(result["record"])
    return result


def process_upload(
    *,
    employee_id: str,
    employee_name: str,
    files: list[tuple[bytes, str]],
    meal_date: str = "",
    overwrite: bool = False,
    source: str = "web_upload",
    manual_amount: Optional[float] = None,
) -> dict:
    """保存图片（最多 3 张）+ 分别识别 + 金额合计入库；同日合并时按内容去重。"""
    if not employee_id:
        return {"ok": False, "status": "error", "error": "缺少员工标识"}
    if not files:
        return {"ok": False, "status": "error", "error": "请至少上传一张截图"}
    if len(files) > MAX_UPLOAD_IMAGES:
        return {
            "ok": False,
            "status": "error",
            "error": f"单次最多上传 {MAX_UPLOAD_IMAGES} 张截图",
        }

    skipped_duplicates = 0
    files, batch_skipped = dedupe_upload_files(files)
    skipped_duplicates += batch_skipped
    if not files:
        return {
            "ok": False,
            "status": "error",
            "error": "所选图片内容重复，请换一张或删除重复文件后重试",
            "skipped_duplicates": skipped_duplicates,
        }

    md_hint = (meal_date or "").strip()
    existing: Optional[dict] = None
    if overwrite and md_hint:
        existing = find_existing_receipt(employee_id, employee_name, md_hint)

    if overwrite and existing:
        files, ext_skipped = filter_files_not_in_hashes(
            files, image_hashes_from_record(existing)
        )
        skipped_duplicates += ext_skipped
        slots = MAX_UPLOAD_IMAGES - len(image_urls_for_record(existing))
        if slots <= 0:
            return {
                "ok": True,
                "status": "unchanged",
                "record": _norm_record(existing),
                "skipped_duplicates": skipped_duplicates,
                "message": f"当日已有 {MAX_UPLOAD_IMAGES} 张凭证，未再添加",
            }
        files = files[:slots]
        if not files:
            return {
                "ok": True,
                "status": "unchanged",
                "record": _norm_record(existing),
                "skipped_duplicates": skipped_duplicates,
                "message": "所选图片与当日已有凭证重复，未重复添加",
            }

    saved: list[tuple[str, str]] = []
    recognitions: list[dict] = []
    for content, filename in files:
        img_path, img_url = save_receipt_image(employee_id, content, filename)
        saved.append((img_path, img_url))
        recognitions.append(recognize_receipt(img_path))

    def _prior_from(rec: Optional[dict]) -> list[dict]:
        if not rec:
            return []
        return recognitions_from_raw_ocr(str(rec.get("raw_ocr") or ""))

    def _recognitions_for_new_only(
        rec: Optional[dict],
    ) -> list[dict]:
        if not rec:
            return recognitions
        seen = image_hashes_from_record(rec)
        out: list[dict] = []
        for (img_path, _url), recog in zip(saved, recognitions):
            try:
                digest = file_content_sha256(Path(img_path).read_bytes())
            except Exception:
                digest = ""
            if digest and digest in seen:
                continue
            out.append(recog)
        return out

    prior_recs = _prior_from(existing) if overwrite and existing else []
    new_recs = _recognitions_for_new_only(existing) if overwrite and existing else recognitions
    combined_for_agg = prior_recs + new_recs

    manual_amt = _norm_amount(manual_amount)
    ok_recs = [r for r in combined_for_agg if r.get("ok")]
    if not ok_recs and manual_amt is None:
        err = next(
            (r.get("error") for r in recognitions if r.get("error")),
            "识别失败",
        )
        return {
            "ok": False,
            "status": "error",
            "error": err,
            "image_urls": [u for _, u in saved],
            "skipped_duplicates": skipped_duplicates,
            "recognized": {
                "ok": False,
                "error": err,
                "images": _per_image_rows(recognitions),
                "image_urls": [u for _, u in saved],
                "image_count": len(saved),
            },
        }

    date_override = md_hint or (
        str(existing.get("meal_date") or "") if existing else ""
    )
    if not ok_recs and manual_amt is not None:
        today = date.today().strftime("%Y-%m-%d")
        amount = manual_amt
        currency = "CNY"
        md = (date_override or today).strip() or today
        merchant = ""
        pending = True
        review_note = "手动填写金额"
        payload = {
            "images": _per_image_rows(recognitions),
            "image_count": len(recognitions),
            "amounts": [manual_amt],
            "manual_amount": True,
        }
    else:
        amount, currency, md, merchant, pending, review_note, payload = aggregate_upload_recognitions(
            combined_for_agg,
            meal_date_override=date_override,
        )
        if amount <= 0 and manual_amt is not None:
            amount = manual_amt
            pending = True
            review_note = "手动填写金额" + (
                f"；{review_note}" if review_note else ""
            )
            payload = dict(payload)
            payload["amounts"] = [manual_amt]
            payload["manual_amount"] = True

    if overwrite and not existing:
        existing = find_existing_receipt(employee_id, employee_name, md)
        if existing:
            prior_recs = _prior_from(existing)
            new_recs = _recognitions_for_new_only(existing)
            combined_for_agg = prior_recs + new_recs
            amount, currency, md, merchant, pending, review_note, payload = (
                aggregate_upload_recognitions(
                    combined_for_agg,
                    meal_date_override=date_override or md,
                )
            )

    if overwrite and existing:
        image_urls_list = merge_saved_images_into_record(existing, saved)
    else:
        image_urls_list = merge_image_url_lists(None, [u for _, u in saved])
    image_url = image_urls_list[0] if image_urls_list else ""
    image_urls_json = json.dumps(image_urls_list, ensure_ascii=False)
    raw_ocr = json.dumps(payload, ensure_ascii=False)

    recognized = {
        "ok": True,
        "date": md,
        "amount": amount,
        "currency": currency,
        "merchant": merchant,
        "image_url": image_url,
        "image_urls": image_urls_list,
        "image_count": len(image_urls_list),
        "amounts": payload.get("amounts", []),
        "pending_review": pending,
        "review_note": review_note,
        "images": payload.get("images", []),
        "skipped_duplicates": skipped_duplicates,
    }

    ok, status, record = upsert_receipt(
        employee_id=employee_id,
        employee_name=employee_name,
        meal_date=md,
        amount=amount,
        currency=currency,
        merchant=merchant,
        image_url=image_url,
        image_urls=image_urls_json,
        raw_ocr=raw_ocr,
        source=source,
        overwrite=overwrite,
        pending_review=pending,
        review_note=review_note,
    )
    if not ok and status == "exists":
        return {
            "ok": False,
            "status": "exists",
            "record": _norm_record(record),
            "recognized": recognized,
            "skipped_duplicates": skipped_duplicates,
        }
    out: dict = {
        "ok": True,
        "status": status,
        "record": _norm_record(record),
        "recognized": recognized,
        "skipped_duplicates": skipped_duplicates,
    }
    if skipped_duplicates > 0:
        out["message"] = f"已跳过 {skipped_duplicates} 张重复图片"
    return out
